import csv
import json
import io
from sqlalchemy.orm import Session
import models

def flatten_dict(d, parent_key='', sep='_'):
    """Flatten a nested dictionary for CSV export."""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            items.append((new_key, json.dumps(v)))
        else:
            items.append((new_key, v))
    return dict(items)

def export_project_data(db: Session, project_id: int, format_type: str):
    """
    Export all approved extracted records for a project in the specified format.
    format_type: 'json', 'csv'
    """
    docs = db.query(models.Document).filter(models.Document.project_id == project_id).all()
    doc_ids = [d.id for d in docs]
    
    records = db.query(models.ExtractedRecord).filter(
        models.ExtractedRecord.document_id.in_(doc_ids),
        models.ExtractedRecord.status == 'approved' # Only export approved records
    ).all()

    return _format_records(records, format_type)

def export_all_approved(db: Session, format_type: str):
    """
    Export ALL approved records across the entire platform.
    """
    records = db.query(models.ExtractedRecord).filter(
        models.ExtractedRecord.status == 'approved'
    ).all()

    return _format_records(records, format_type)

def export_by_document(db: Session, document_id: int, format_type: str):
    """
    Export records for a single document (approved or pending review).
    """
    records = db.query(models.ExtractedRecord).filter(
        models.ExtractedRecord.document_id == document_id
    ).all()

    return _format_records(records, format_type)

def _format_records(records, format_type: str):
    """
    Core formatting logic shared by all export routes.
    """
    if not records:
        return None

    if format_type == 'json':
        data = []
        for r in records:
            data.append({
                "record_id": r.id,
                "document_id": r.document_id,
                "confidence": r.confidence,
                "data": r.record_data
            })
        return {"data": data, "type": "application/json", "filename": "export.json"}
        
    elif format_type in ['csv', 'xlsx']:
        flattened = []
        all_keys = set()
        
        for r in records:
            data = r.record_data
            
            # Extract root fields and list fields
            root_fields = {}
            list_fields = {}
            if isinstance(data, dict):
                for k, v in data.items():
                    if 'field_confidences' in k:
                        continue
                    if isinstance(v, list):
                        list_fields[k] = v
                    else:
                        if isinstance(v, dict):
                            root_fields.update(flatten_dict(v, k))
                        else:
                            root_fields[k] = v
                            
            if list_fields:
                # Unroll list fields (e.g. 'data' array) and repeat root fields for each row
                for list_key, list_val in list_fields.items():
                    for idx, item in enumerate(list_val):
                        flat = dict(root_fields)
                        if isinstance(item, dict):
                            flat.update(flatten_dict(item, list_key))
                        else:
                            flat[list_key] = item
                            
                        flat['_record_id'] = f"{r.id}_{idx+1}"
                        flat['_document_id'] = r.document_id
                        flat['_confidence'] = r.confidence
                        all_keys.update(flat.keys())
                        flattened.append(flat)
            elif isinstance(data, list):
                # Legacy array-only schema
                for idx, item in enumerate(data):
                    flat = flatten_dict(item) if isinstance(item, dict) else {"value": item}
                    flat['_record_id'] = f"{r.id}_{idx+1}"
                    flat['_document_id'] = r.document_id
                    flat['_confidence'] = r.confidence
                    all_keys.update(flat.keys())
                    flattened.append(flat)
            else:
                # Simple object without lists
                flat = flatten_dict(data) if isinstance(data, dict) else {"value": data}
                keys_to_remove = [k for k in flat.keys() if 'field_confidences' in k]
                for k in keys_to_remove:
                    del flat[k]
                flat['_record_id'] = r.id
                flat['_document_id'] = r.document_id
                flat['_confidence'] = r.confidence
                all_keys.update(flat.keys())
                flattened.append(flat)
                
        # Sort keys
        meta_keys = sorted([k for k in all_keys if k.startswith('_')])
        data_keys = sorted([k for k in all_keys if not k.startswith('_')])
        ordered_keys = meta_keys + data_keys
        
        if format_type == 'csv':
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=ordered_keys, extrasaction='ignore')
            writer.writeheader()
            for row in flattened:
                writer.writerow(row)
            return {"data": output.getvalue(), "type": "text/csv", "filename": "export.csv"}
            
        elif format_type == 'xlsx':
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Exported Data"
            
            ws.append(ordered_keys)
            
            header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            header_font = Font(bold=True)
            for col_idx in range(1, len(ordered_keys) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                
            for row in flattened:
                row_data = [row.get(key, "") for key in ordered_keys]
                ws.append(row_data)
                
            for col_idx, key in enumerate(ordered_keys, 1):
                col_letter = get_column_letter(col_idx)
                max_len = len(str(key))
                is_number_col = any(keyword in key.lower() for keyword in ['amount', 'price', 'quantity', 'total', 'tax'])
                
                for row_idx in range(2, len(flattened) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell_val = str(cell.value) if cell.value is not None else ""
                    if len(cell_val) > max_len:
                        max_len = len(cell_val)
                    
                    if is_number_col and cell.value:
                        try:
                            cell.value = float(cell.value)
                            cell.number_format = '#,##0.00'
                        except ValueError:
                            pass
                            
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
                
            ws.page_setup.fitToWidth = 1
            
            output = io.BytesIO()
            wb.save(output)
            wb.close()
            return {"data": output.getvalue(), "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "filename": "export.xlsx"}

        
    elif format_type == 'txt':
        output = io.StringIO()
        for i, r in enumerate(records):
            if i > 0:
                output.write("\n\n" + "="*50 + "\n\n")
            output.write(f"--- Document ID {r.document_id} ---\n\n")
            if isinstance(r.record_data, dict) and "full_transcription" in r.record_data:
                output.write(r.record_data["full_transcription"])
            else:
                output.write(json.dumps(r.record_data, indent=2))
                
        return {"data": output.getvalue(), "type": "text/plain", "filename": "export.txt"}
        
    elif format_type == 'pdf':
        try:
            from pypdf import PdfWriter
            import os
            
            merger = PdfWriter()
            has_pdfs = False
            
            for r in records:
                # Need to find the actual Document path. We can import Session and get it, or assume record.document
                # Since records are queried, let's rely on the router passing db if needed, or query it here.
                # Actually, `r` has a relationship `r.document` if defined in SQLAlchemy models.
                doc = r.document
                pdf_path = doc.file_path
                if os.path.exists(pdf_path) and pdf_path.lower().endswith('.pdf'):
                    merger.append(pdf_path)
                    has_pdfs = True
            
            if not has_pdfs:
                return None
                
            output = io.BytesIO()
            merger.write(output)
            merger.close()
            
            return {"data": output.getvalue(), "type": "application/pdf", "filename": "export.pdf"}
        except Exception as exc:
            print(f"Error generating PDF export: {exc}")
            return None

    return None
